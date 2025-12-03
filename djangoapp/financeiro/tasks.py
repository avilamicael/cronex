# financeiro/tasks.py
from celery import shared_task
from django.utils.timezone import now
from datetime import timedelta
from django.contrib.auth import get_user_model
from django.db.models import Sum, Count
from financeiro.models import ContaPagar, RelatorioFaturamentoMensal, Filial
from core.notificacoes import enviar_mensagem_telegram
from collections import defaultdict
from core.utils import formatar_brl
from accounts.models import Empresa
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
import tempfile
import zipfile
import shutil
from django.core.files import File
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook

User = get_user_model()

@shared_task(name="Verificar contas vencidas")
def notificar_contas_vencidas():
    hoje = now().date()

    for user in User.objects.filter(telegram_chat_id__isnull=False, ativo=True):
        contas = ContaPagar.objects.filter(
            empresa=user.empresa,
            status='vencida',
            valor_pago__lte=0
        )

        if not contas.exists():
            continue

        # Agrupamento por filial e fornecedor
        dados_agrupados = (
            contas
            .values('filial__nome', 'fornecedor__nome')
            .annotate(
                total=Sum('valor_bruto'),
                quantidade=Count('id')
            )
            .order_by('filial__nome', 'fornecedor__nome')
        )

        # Organiza em um dicionário por filial
        por_filial = defaultdict(list)
        for item in dados_agrupados:
            filial = item['filial__nome'] or "Filial não informada"
            fornecedor = item['fornecedor__nome'] or "Fornecedor não informado"
            total = item['total'] or 0
            quantidade = item['quantidade']
            por_filial[filial].append((fornecedor, total, quantidade))

        # Monta a mensagem
        mensagem = "━━━━━━━━━━━━━━━━━━━━━━\n"
        mensagem += f"📅 <b>{hoje.strftime('%d/%m/%Y')}</b>\n"
        mensagem += "<b>Contas Vencidas</b>\n"        
        mensagem += "━━━━━━━━━━━━━━━━━━━━━━\n"
        for filial, fornecedores in por_filial.items():
            mensagem += f"\n<b>{filial}</b>:\n"
            total_filial = 0
            qtd_filial = 0
            for fornecedor, total, quantidade in fornecedores:
                mensagem += f"• {fornecedor} - {formatar_brl(total)} - {quantidade} conta(s)\n"
                total_filial += total
                qtd_filial += quantidade
            mensagem += f"➡️ <b>Total: {formatar_brl(total_filial)} ({qtd_filial} conta(s))</b>\n"


        total_geral = contas.aggregate(total=Sum('valor_bruto'))['total'] or 0
        mensagem += f"\n<b>Total geral: {contas.count()} conta(s)</b>"
        mensagem += f"\n<b>Valor total: {formatar_brl(total_geral)}</b>"

        enviar_mensagem_telegram(user.telegram_chat_id, mensagem)

@shared_task(name="Verificar contas a vencer nos próximos 7 dias")
def notificar_contas_a_vencer():
    hoje = now().date()
    limite = hoje + timedelta(days=7)

    for user in User.objects.filter(telegram_chat_id__isnull=False, ativo=True):
        contas = ContaPagar.objects.filter(
            empresa=user.empresa,
            status='a_vencer',
            data_vencimento__gte=hoje,
            data_vencimento__lte=limite
        )

        if not contas.exists():
            continue

        # Agrupar por filial e fornecedor
        dados_agrupados = (
            contas
            .values('filial__nome', 'fornecedor__nome')
            .annotate(
                total=Sum('valor_bruto'),
                quantidade=Count('id')
            )
            .order_by('filial__nome', 'fornecedor__nome')
        )

        por_filial = defaultdict(list)
        for item in dados_agrupados:
            filial = item['filial__nome'] or "Filial não informada"
            fornecedor = item['fornecedor__nome'] or "Fornecedor não informado"
            total = item['total'] or 0
            quantidade = item['quantidade']
            por_filial[filial].append((fornecedor, total, quantidade))

        # Montar a mensagem
        mensagem = "━━━━━━━━━━━━━━━━━━━━━━\n"
        mensagem += f"📅 <b>{hoje.strftime('%d/%m/%Y')}</b>\n"
        mensagem += "<b>Contas a vencer nos próximos 7 dias</b>\n"
        mensagem += "━━━━━━━━━━━━━━━━━━━━━━\n"

        for filial, fornecedores in por_filial.items():
            mensagem += f"\n<b>{filial}</b>:\n"
            total_filial = 0
            qtd_filial = 0
            for fornecedor, total, quantidade in fornecedores:
                mensagem += f"• {fornecedor} - {formatar_brl(total)} - {quantidade} conta(s)\n"
                total_filial += total
                qtd_filial += quantidade
            mensagem += f"➡️ <b>Total: {formatar_brl(total_filial)} ({qtd_filial} conta(s))</b>\n"

        total_geral = contas.aggregate(total=Sum('valor_bruto'))['total'] or 0
        mensagem += f"\n<b>Total geral: {contas.count()} conta(s)</b>"
        mensagem += f"\n<b>Valor total: {formatar_brl(total_geral)}</b>"

        enviar_mensagem_telegram(user.telegram_chat_id, mensagem)

@shared_task(name="Atualizar status de contas vencidas")
def atualizar_status_contas():
    hoje = now().date()
    contas_a_vencer = ContaPagar.objects.filter(status='a_vencer', data_vencimento__lt=hoje)
    atualizadas = contas_a_vencer.update(status='vencida')
    return f"{atualizadas} contas atualizadas para 'vencida'"

def gerar_excel_filial(filial, contas, mes, ano):
    """Gera um arquivo Excel com as contas pagas de uma filial"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{filial.nome[:30]}"  # Limita o tamanho do nome da aba

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Título
    ws.merge_cells('A1:M1')
    titulo_cell = ws['A1']
    titulo_cell.value = f"RELATÓRIO DE FATURAMENTO - {filial.nome}"
    titulo_cell.font = Font(bold=True, size=14)
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Subtítulo com período
    ws.merge_cells('A2:M2')
    subtitulo_cell = ws['A2']
    meses = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
             'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    subtitulo_cell.value = f"Período: {meses[mes]}/{ano}"
    subtitulo_cell.font = Font(bold=True, size=11)
    subtitulo_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Linha em branco
    ws.append([])

    # Cabeçalhos
    headers = [
        'Filial', 'Banco Pagamento', 'Transação', 'Fornecedor', 'Documento',
        'Data Movimentação', 'Data Vencimento', 'Data Pagamento',
        'Valor Bruto', 'Juros', 'Multa', 'Valor Pago', 'Nº Notas'
    ]
    ws.append(headers)

    # Estiliza os cabeçalhos
    header_row = ws[4]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Dados
    total_valor_bruto = 0
    total_valor_pago = 0

    for conta in contas:
        ws.append([
            conta.filial.nome,
            conta.conta_bancaria_pagamento.nome if conta.conta_bancaria_pagamento else 'NÃO INFORMADO',
            conta.transacao.nome,
            conta.fornecedor.nome if conta.fornecedor else '',
            conta.documento,
            conta.data_movimentacao.strftime('%d/%m/%Y'),
            conta.data_vencimento.strftime('%d/%m/%Y'),
            conta.data_pagamento.strftime('%d/%m/%Y') if conta.data_pagamento else '',
            float(conta.valor_bruto),
            float(conta.valor_juros),
            float(conta.valor_multa),
            float(conta.valor_pago),
            conta.numero_notas,
        ])
        total_valor_bruto += conta.valor_bruto
        total_valor_pago += conta.valor_pago

    # Linha de totais
    ultima_linha = ws.max_row + 1
    ws.append([
        '', '', '', '', '', '', '', 'TOTAL:',
        float(total_valor_bruto), '', '', float(total_valor_pago), ''
    ])

    # Estiliza a linha de totais
    total_row = ws[ultima_linha]
    for cell in total_row:
        cell.font = Font(bold=True)
        if cell.column in [9, 12]:  # Colunas de valores
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Ajusta largura das colunas
    column_widths = {
        'A': 20, 'B': 20, 'C': 20, 'D': 30, 'E': 15,
        'F': 18, 'G': 18, 'H': 18, 'I': 15, 'J': 10,
        'K': 10, 'L': 15, 'M': 20
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    return wb

@shared_task(name="Gerar relatório de faturamento mensal")
def gerar_relatorio_faturamento_mensal(mes_ref=None, ano_ref=None):
    """
    Task que roda todo dia 1 de cada mês para gerar o relatório do mês anterior
    Agrupa as contas por CONTA BANCÁRIA (campo conta_bancaria da Filial)

    Args:
        mes_ref: Mês de referência (1-12). Se None, usa mês anterior
        ano_ref: Ano de referência. Se None, usa ano do mês anterior
    """
    hoje = now().date()

    if mes_ref and ano_ref:
        mes = mes_ref
        ano = ano_ref
        print(f"[Relatório] Modo manual: gerando para {mes:02d}/{ano}")
    else:
        mes_anterior = hoje - relativedelta(months=1)
        mes = mes_anterior.month
        ano = mes_anterior.year
        print(f"[Relatório] Modo automático: hoje={hoje}, gerando para {mes:02d}/{ano}")

    empresas = Empresa.objects.filter(ativo=True)
    print(f"[Relatório] Empresas ativas encontradas: {empresas.count()}")

    for empresa in empresas:
        try:
            print(f"\n[Relatório] Processando empresa: {empresa.nome} (ID={empresa.id})")

            # Busca contas pagas no mês anterior
            contas = ContaPagar.objects.filter(
                empresa=empresa,
                status='pago',
                data_pagamento__month=mes,
                data_pagamento__year=ano
            ).select_related('filial', 'conta_bancaria_pagamento', 'transacao', 'fornecedor')

            total_contas = contas.count()
            print(f"[Relatório] {empresa.nome}: {total_contas} contas pagas em {mes:02d}/{ano}")

            if not contas.exists():
                print(f"[Relatório] {empresa.nome}: Nenhuma conta paga, pulando geração")
                continue

            if total_contas > 0:
                print(f"[Relatório] Primeiras contas para debug:")
                for c in contas[:3]:
                    # conta_bancaria_pagamento é uma Filial
                    if c.conta_bancaria_pagamento:
                        banco_info = f"{c.conta_bancaria_pagamento.nome} - Conta: {c.conta_bancaria_pagamento.conta_bancaria or 'N/A'}"
                    else:
                        banco_info = 'SEM BANCO'
                    print(f"  - ID={c.id}, Banco={banco_info}, Filial origem={c.filial.nome if c.filial else 'SEM FILIAL'}, Valor={c.valor_pago}")

            # Agrupa por CONTA BANCÁRIA (Filial que é o banco pagador)
            bancos_com_contas = {}
            contas_sem_banco = []
            
            for conta in contas:
                banco_pagador = conta.conta_bancaria_pagamento  # É uma Filial
                
                if banco_pagador:
                    # Usa a Filial pagadora como chave
                    if banco_pagador not in bancos_com_contas:
                        bancos_com_contas[banco_pagador] = []
                    bancos_com_contas[banco_pagador].append(conta)
                else:
                    contas_sem_banco.append(conta)

            print(f"[Relatório] Contas bancárias (Filiais pagadoras) encontradas: {len(bancos_com_contas)}")
            for banco_pagador, contas_banco in bancos_com_contas.items():
                conta_info = banco_pagador.conta_bancaria or 'Conta não cadastrada'
                print(f"  - {banco_pagador.nome} (Conta: {conta_info}): {len(contas_banco)} contas")
            
            if contas_sem_banco:
                print(f"  - ⚠️ ATENÇÃO: {len(contas_sem_banco)} contas SEM BANCO associado")

            # Valida se tem algo para gerar
            if not bancos_com_contas and not contas_sem_banco:
                print(f"[Relatório] ⚠️ Nenhuma conta encontrada para processar!")
                continue

            # Cria diretório temporário
            temp_dir = tempfile.mkdtemp()
            print(f"[Relatório] Diretório temporário: {temp_dir}")
            
            try:
                # Gera Excel para cada CONTA BANCÁRIA (Filial pagadora)
                for banco_pagador, contas_banco in bancos_com_contas.items():
                    try:
                        # Nome seguro para pasta usando o nome da Filial pagadora
                        nome_pasta = banco_pagador.nome.replace('/', '-').replace('\\', '-').replace(' ', '_')
                        pasta_banco = os.path.join(temp_dir, nome_pasta)
                        os.makedirs(pasta_banco, exist_ok=True)
                        
                        conta_info = banco_pagador.conta_bancaria or 'N/A'
                        print(f"[Relatório] Gerando Excel para: {banco_pagador.nome} (Conta: {conta_info})")
                        print(f"[Relatório]   Pasta: {pasta_banco}")
                        print(f"[Relatório]   Total de contas: {len(contas_banco)}")

                        # Gera o Excel
                        wb = gerar_excel_banco(banco_pagador, contas_banco, mes, ano)

                        # Salva
                        arquivo_excel = os.path.join(pasta_banco, f'faturamento_{mes:02d}_{ano}.xlsx')
                        wb.save(arquivo_excel)
                        print(f"[Relatório]   ✅ Excel salvo: {arquivo_excel}")
                        
                    except Exception as e:
                        print(f"[Relatório]   ❌ ERRO ao gerar Excel para {banco_pagador.nome}: {e}")
                        import traceback
                        traceback.print_exc()
                        continue

                # Se houver contas sem banco
                if contas_sem_banco:
                    try:
                        nome_pasta = "Sem_Banco"
                        pasta_sem_banco = os.path.join(temp_dir, nome_pasta)
                        os.makedirs(pasta_sem_banco, exist_ok=True)
                        
                        print(f"[Relatório] Gerando Excel para contas SEM BANCO ({len(contas_sem_banco)} contas)")
                        
                        wb = gerar_excel_sem_banco(contas_sem_banco, mes, ano)
                        arquivo_excel = os.path.join(pasta_sem_banco, f'faturamento_{mes:02d}_{ano}.xlsx')
                        wb.save(arquivo_excel)
                        print(f"[Relatório]   ✅ Excel salvo: {arquivo_excel}")
                        
                    except Exception as e:
                        print(f"[Relatório]   ❌ ERRO ao gerar Excel para contas sem banco: {e}")
                        import traceback
                        traceback.print_exc()

                # Lista arquivos gerados
                print(f"\n[Relatório] Listando arquivos gerados:")
                arquivos_excel = []
                for root, dirs, files in os.walk(temp_dir):
                    for file in files:
                        if file.endswith('.xlsx'):
                            file_path = os.path.join(root, file)
                            arquivos_excel.append(file_path)
                            print(f"  - {os.path.relpath(file_path, temp_dir)}")
                
                if not arquivos_excel:
                    print(f"[Relatório] ❌ ERRO: Nenhum arquivo Excel foi gerado!")
                    continue

                # Cria ZIP
                zip_filename = f'relatorio_faturamento_{mes:02d}_{ano}_{empresa.nome.replace(" ", "_")}.zip'
                zip_temp_path = os.path.join(temp_dir, zip_filename)
                print(f"\n[Relatório] Criando ZIP: {zip_filename}")

                with zipfile.ZipFile(zip_temp_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    arquivos_adicionados = 0
                    for file_path in arquivos_excel:
                        arcname = os.path.relpath(file_path, temp_dir)
                        zipf.write(file_path, arcname)
                        arquivos_adicionados += 1
                        print(f"  - Adicionado: {arcname}")

                print(f"[Relatório] ✅ ZIP criado com {arquivos_adicionados} arquivos")

                # Verifica se o ZIP foi criado
                if not os.path.exists(zip_temp_path):
                    print(f"[Relatório] ❌ ERRO: ZIP não foi criado em {zip_temp_path}")
                    continue
                
                zip_size = os.path.getsize(zip_temp_path)
                print(f"[Relatório] Tamanho do ZIP: {zip_size / 1024:.2f} KB")

                # Salva no banco
                relatorio, created = RelatorioFaturamentoMensal.objects.update_or_create(
                    empresa=empresa,
                    mes=mes,
                    ano=ano,
                    defaults={'gerado_por': None}
                )
                print(f"[Relatório] Registro no DB: {'CRIADO' if created else 'ATUALIZADO'} (ID={relatorio.id})")

                if relatorio.arquivo_zip:
                    print(f"[Relatório] Arquivo antigo será substituído: {relatorio.arquivo_zip.name}")

                # Salva arquivo
                print(f"[Relatório] Salvando arquivo no Django...")
                with open(zip_temp_path, 'rb') as f:
                    relatorio.arquivo_zip.save(zip_filename, File(f), save=True)

                print(f"[Relatório] ✅ Arquivo salvo: {relatorio.arquivo_zip.name}")
                if hasattr(relatorio.arquivo_zip, 'path'):
                    print(f"[Relatório] Path completo: {relatorio.arquivo_zip.path}")
                
                print(f"\n[Relatório] ✅✅✅ SUCESSO para {empresa.nome}!")

            except Exception as e:
                print(f"[Relatório] ❌ ERRO durante geração de arquivos: {e}")
                import traceback
                traceback.print_exc()
                
            finally:
                # Limpa temp
                print(f"[Relatório] Limpando diretório temporário...")
                shutil.rmtree(temp_dir, ignore_errors=True)

        except Exception as e:
            print(f"\n[Relatório] ❌❌❌ ERRO GERAL para {empresa.nome}: {e}")
            import traceback
            traceback.print_exc()
            continue

    print(f"\n[Relatório] ========== PROCESSO FINALIZADO ==========")
    return f"Relatórios gerados para {mes:02d}/{ano}"


def gerar_excel_banco(banco_pagador, contas, mes, ano):
    """
    Gera Excel com contas pagas por uma conta bancária (Filial pagadora)
    
    Args:
        banco_pagador: Filial que representa a conta bancária pagadora
        contas: Lista de ContaPagar pagas por esta conta
        mes: Mês de referência
        ano: Ano de referência
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    print(f"    [Excel] Iniciando geração para: {banco_pagador.nome}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Faturamento"
    
    # Cabeçalho
    ws['A1'] = f'Relatório de Faturamento - {mes:02d}/{ano}'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f'Conta Bancária Pagadora: {banco_pagador.nome}'
    ws['A2'].font = Font(size=12, bold=True)
    
    # Headers - Todas as colunas do relatório original + Banco Pagamento
    headers = [
        'Data Movim.',
        'Data Venc.',
        'Data Pgto',
        'Filial',
        'Banco Pagamento',
        'Transação',
        'Fornecedor',
        'Tipo Pgto',
        'Documento',
        'Descrição',
        'Nº Notas',
        'Valor Bruto',
        'Juros',
        'Multa',
        'Acréscimos',
        'Desconto',
        'Valor Pago',
        'Saldo'
    ]
    
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Dados
    row = 5
    total_bruto = 0
    total_juros = 0
    total_multa = 0
    total_acrescimos = 0
    total_desconto = 0
    total_pago = 0
    total_saldo = 0
    
    print(f"    [Excel] Processando {len(contas)} contas...")
    
    for conta in contas:
        col = 1
        
        # Data Movimentação
        ws.cell(row=row, column=col).value = conta.data_movimentacao.strftime('%d/%m/%Y') if conta.data_movimentacao else '-'
        col += 1
        
        # Data Vencimento
        ws.cell(row=row, column=col).value = conta.data_vencimento.strftime('%d/%m/%Y') if conta.data_vencimento else '-'
        col += 1
        
        # Data Pagamento
        ws.cell(row=row, column=col).value = conta.data_pagamento.strftime('%d/%m/%Y') if conta.data_pagamento else '-'
        col += 1
        
        # Filial (origem da despesa)
        ws.cell(row=row, column=col).value = conta.filial.nome if conta.filial else '-'
        col += 1
        
        # Banco Pagamento (conta bancária que efetuou o pagamento)
        ws.cell(row=row, column=col).value = conta.conta_bancaria_pagamento.nome if conta.conta_bancaria_pagamento else '-'
        col += 1
        
        # Transação
        ws.cell(row=row, column=col).value = conta.transacao.nome if conta.transacao else '-'
        col += 1
        
        # Fornecedor
        ws.cell(row=row, column=col).value = conta.fornecedor.nome if conta.fornecedor else '-'
        col += 1
        
        # Tipo Pagamento
        ws.cell(row=row, column=col).value = conta.tipo_pagamento.nome if conta.tipo_pagamento else '-'
        col += 1
        
        # Documento
        ws.cell(row=row, column=col).value = conta.documento or '-'
        col += 1
        
        # Descrição
        ws.cell(row=row, column=col).value = conta.descricao or '-'
        col += 1
        
        # Número Notas
        ws.cell(row=row, column=col).value = conta.numero_notas or '-'
        col += 1
        
        # Valor Bruto
        valor_bruto = float(conta.valor_bruto) if conta.valor_bruto else 0
        ws.cell(row=row, column=col).value = valor_bruto
        ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'
        total_bruto += valor_bruto
        col += 1
        
        # Juros
        valor_juros = float(conta.valor_juros) if conta.valor_juros else 0
        ws.cell(row=row, column=col).value = valor_juros
        ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'
        total_juros += valor_juros
        col += 1
        
        # Multa
        valor_multa = float(conta.valor_multa) if conta.valor_multa else 0
        ws.cell(row=row, column=col).value = valor_multa
        ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'
        total_multa += valor_multa
        col += 1
        
        # Acréscimos
        valor_acrescimos = float(conta.outros_acrescimos) if conta.outros_acrescimos else 0
        ws.cell(row=row, column=col).value = valor_acrescimos
        ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'
        total_acrescimos += valor_acrescimos
        col += 1
        
        # Desconto
        valor_desconto = float(conta.valor_desconto) if conta.valor_desconto else 0
        ws.cell(row=row, column=col).value = valor_desconto
        ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'
        total_desconto += valor_desconto
        col += 1
        
        # Valor Pago
        valor_pago = float(conta.valor_pago) if conta.valor_pago else 0
        ws.cell(row=row, column=col).value = valor_pago
        ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'
        total_pago += valor_pago
        col += 1
        
        # Saldo
        valor_saldo = float(conta.valor_saldo) if conta.valor_saldo else 0
        ws.cell(row=row, column=col).value = valor_saldo
        ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'
        total_saldo += valor_saldo
        
        row += 1
    
    # Linha de Totais
    ws.cell(row=row, column=11).value = 'TOTAL:'
    ws.cell(row=row, column=11).font = Font(bold=True)
    ws.cell(row=row, column=11).alignment = Alignment(horizontal='right')
    
    ws.cell(row=row, column=12).value = total_bruto
    ws.cell(row=row, column=12).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=12).font = Font(bold=True)
    
    ws.cell(row=row, column=13).value = total_juros
    ws.cell(row=row, column=13).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=13).font = Font(bold=True)
    
    ws.cell(row=row, column=14).value = total_multa
    ws.cell(row=row, column=14).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=14).font = Font(bold=True)
    
    ws.cell(row=row, column=15).value = total_acrescimos
    ws.cell(row=row, column=15).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=15).font = Font(bold=True)
    
    ws.cell(row=row, column=16).value = total_desconto
    ws.cell(row=row, column=16).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=16).font = Font(bold=True)
    
    ws.cell(row=row, column=17).value = total_pago
    ws.cell(row=row, column=17).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=17).font = Font(bold=True)
    
    ws.cell(row=row, column=18).value = total_saldo
    ws.cell(row=row, column=18).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=18).font = Font(bold=True)
    
    # Ajusta larguras das colunas
    ws.column_dimensions['A'].width = 12  # Data Movim
    ws.column_dimensions['B'].width = 12  # Data Venc
    ws.column_dimensions['C'].width = 12  # Data Pgto
    ws.column_dimensions['D'].width = 25  # Filial
    ws.column_dimensions['E'].width = 25  # Banco Pagamento
    ws.column_dimensions['F'].width = 25  # Transação
    ws.column_dimensions['G'].width = 30  # Fornecedor
    ws.column_dimensions['H'].width = 18  # Tipo Pgto
    ws.column_dimensions['I'].width = 20  # Documento
    ws.column_dimensions['J'].width = 40  # Descrição
    ws.column_dimensions['K'].width = 15  # Nº Notas
    ws.column_dimensions['L'].width = 14  # Valor Bruto
    ws.column_dimensions['M'].width = 12  # Juros
    ws.column_dimensions['N'].width = 12  # Multa
    ws.column_dimensions['O'].width = 12  # Acréscimos
    ws.column_dimensions['P'].width = 12  # Desconto
    ws.column_dimensions['Q'].width = 14  # Valor Pago
    ws.column_dimensions['R'].width = 12  # Saldo
    
    print(f"    [Excel] ✅ Excel gerado com {len(contas)} linhas, total pago: R$ {total_pago:,.2f}")
    
    return wb


def gerar_excel_sem_banco(contas, mes, ano):
    """Gera Excel para contas sem banco associado"""    
    print(f"    [Excel] Iniciando geração para contas SEM BANCO")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Faturamento"
    
    # Cabeçalho
    ws['A1'] = f'Relatório de Faturamento - {mes:02d}/{ano}'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = '⚠️ CONTAS SEM BANCO ASSOCIADO ⚠️'
    ws['A2'].font = Font(size=12, bold=True, color='FF0000')
    
    # Headers - Todas as colunas
    headers = [
        'Data Movim.',
        'Data Venc.',
        'Data Pgto',
        'Filial',
        'Banco Pagamento',
        'Transação',
        'Fornecedor',
        'Tipo Pgto',
        'Documento',
        'Descrição',
        'Nº Notas',
        'Valor Bruto',
        'Juros',
        'Multa',
        'Acréscimos',
        'Desconto',
        'Valor Pago',
        'Saldo'
    ]
    
    header_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Dados
    row = 5
    total_bruto = 0
    total_juros = 0
    total_multa = 0
    total_acrescimos = 0
    total_desconto = 0
    total_pago = 0
    total_saldo = 0
    
    for conta in contas:
        col = 1
        
        # Data Movimentação
        ws.cell(row=row, column=col).value = conta.data_movimentacao.strftime('%d/%m/%Y') if conta.data_movimentacao else '-'
        col += 1
        
        # Data Vencimento
        ws.cell(row=row, column=col).value = conta.data_vencimento.strftime('%d/%m/%Y') if conta.data_vencimento else '-'
        col += 1
        
        # Data Pagamento
        ws.cell(row=row, column=col).value = conta.data_pagamento.strftime('%d/%m/%Y') if conta.data_pagamento else '-'
        col += 1
        
        # Filial
        ws.cell(row=row, column=col).value = conta.filial.nome if conta.filial else '-'
        col += 1
        
        # Banco Pagamento (sempre mostra como "SEM BANCO" neste relatório)
        ws.cell(row=row, column=col).value = 'SEM BANCO'
        ws.cell(row=row, column=col).font = Font(color='FF0000')
        col += 1
        
        # Transação
        ws.cell(row=row, column=col).value = conta.transacao.nome if conta.transacao else '-'
        col += 1
        
        # Fornecedor
        ws.cell(row=row, column=col).value = conta.fornecedor.nome if conta.fornecedor else '-'
        col += 1
        
        # Tipo Pagamento
        ws.cell(row=row, column=col).value = conta.tipo_pagamento.nome if conta.tipo_pagamento else '-'
        col += 1
        
        # Documento
        ws.cell(row=row, column=col).value = conta.documento or '-'
        col += 1
        
        # Descrição
        ws.cell(row=row, column=col).value = conta.descricao or '-'
        col += 1
        
        # Número Notas
        ws.cell(row=row, column=col).value = conta.numero_notas or '-'
        col += 1
        
        # Valor Bruto
        valor_bruto = float(conta.valor_bruto) if conta.valor_bruto else 0
        ws.cell(row=row, column=col).value = valor_bruto
        ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'
        total_bruto += valor_bruto
        col += 1
        
        # Juros
        valor_juros = float(conta.valor_juros) if conta.valor_juros else 0
        ws.cell(row=row, column=col).value = valor_juros
        ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'
        total_juros += valor_juros
        col += 1
        
        # Multa
        valor_multa = float(conta.valor_multa) if conta.valor_multa else 0
        ws.cell(row=row, column=col).value = valor_multa
        ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'
        total_multa += valor_multa
        col += 1
        
        # Acréscimos
        valor_acrescimos = float(conta.outros_acrescimos) if conta.outros_acrescimos else 0
        ws.cell(row=row, column=col).value = valor_acrescimos
        ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'
        total_acrescimos += valor_acrescimos
        col += 1
        
        # Desconto
        valor_desconto = float(conta.valor_desconto) if conta.valor_desconto else 0
        ws.cell(row=row, column=col).value = valor_desconto
        ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'
        total_desconto += valor_desconto
        col += 1
        
        # Valor Pago
        valor_pago = float(conta.valor_pago) if conta.valor_pago else 0
        ws.cell(row=row, column=col).value = valor_pago
        ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'
        total_pago += valor_pago
        col += 1
        
        # Saldo
        valor_saldo = float(conta.valor_saldo) if conta.valor_saldo else 0
        ws.cell(row=row, column=col).value = valor_saldo
        ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'
        total_saldo += valor_saldo
        
        row += 1
    
    # Linha de Totais
    ws.cell(row=row, column=11).value = 'TOTAL:'
    ws.cell(row=row, column=11).font = Font(bold=True)
    ws.cell(row=row, column=11).alignment = Alignment(horizontal='right')
    
    ws.cell(row=row, column=12).value = total_bruto
    ws.cell(row=row, column=12).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=12).font = Font(bold=True)
    
    ws.cell(row=row, column=13).value = total_juros
    ws.cell(row=row, column=13).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=13).font = Font(bold=True)
    
    ws.cell(row=row, column=14).value = total_multa
    ws.cell(row=row, column=14).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=14).font = Font(bold=True)
    
    ws.cell(row=row, column=15).value = total_acrescimos
    ws.cell(row=row, column=15).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=15).font = Font(bold=True)
    
    ws.cell(row=row, column=16).value = total_desconto
    ws.cell(row=row, column=16).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=16).font = Font(bold=True)
    
    ws.cell(row=row, column=17).value = total_pago
    ws.cell(row=row, column=17).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=17).font = Font(bold=True)
    
    ws.cell(row=row, column=18).value = total_saldo
    ws.cell(row=row, column=18).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=18).font = Font(bold=True)
    
    # Ajusta larguras das colunas
    ws.column_dimensions['A'].width = 12  # Data Movim
    ws.column_dimensions['B'].width = 12  # Data Venc
    ws.column_dimensions['C'].width = 12  # Data Pgto
    ws.column_dimensions['D'].width = 25  # Filial
    ws.column_dimensions['E'].width = 25  # Banco Pagamento
    ws.column_dimensions['F'].width = 25  # Transação
    ws.column_dimensions['G'].width = 30  # Fornecedor
    ws.column_dimensions['H'].width = 18  # Tipo Pgto
    ws.column_dimensions['I'].width = 20  # Documento
    ws.column_dimensions['J'].width = 40  # Descrição
    ws.column_dimensions['K'].width = 15  # Nº Notas
    ws.column_dimensions['L'].width = 14  # Valor Bruto
    ws.column_dimensions['M'].width = 12  # Juros
    ws.column_dimensions['N'].width = 12  # Multa
    ws.column_dimensions['O'].width = 12  # Acréscimos
    ws.column_dimensions['P'].width = 12  # Desconto
    ws.column_dimensions['Q'].width = 14  # Valor Pago
    ws.column_dimensions['R'].width = 12  # Saldo
    
    print(f"    [Excel] ✅ Excel SEM BANCO gerado com {len(contas)} linhas, total pago: R$ {total_pago:,.2f}")
    
    return wb

# ========================================
# TASKS PARA IMPORTAÇÃO AUTOMÁTICA DE NFE
# ========================================

@shared_task(name="Buscar notas fiscais automaticamente")
def buscar_notas_automaticamente():
    """
    Task que busca novas notas fiscais na SEFAZ de forma automática.
    Roda a cada 4 horas para todos os certificados com busca automática ativa.
    """
    from financeiro.models import ConfiguracaoNFe, CertificadoDigital, NotaFiscal
    from financeiro.crypto import decrypt_password
    from financeiro.nfe.sefaz_client import SefazClient
    from django.core.files.base import ContentFile
    from django.db import transaction
    from django.utils import timezone

    print(f"[NFe Auto] Iniciando busca automática - {timezone.now()}")

    # Busca todas as configurações ativas
    configs = ConfiguracaoNFe.objects.filter(
        busca_automatica_ativa=True,
        certificado__ativo=True
    ).select_related('certificado', 'certificado__filial', 'certificado__empresa')

    total_configs = configs.count()
    print(f"[NFe Auto] Encontradas {total_configs} configuração(ões) ativa(s)")

    if total_configs == 0:
        print("[NFe Auto] Nenhuma configuração ativa. Finalizando.")
        return "Nenhuma configuração ativa"

    resultados = []

    for config in configs:
        certificado = config.certificado
        empresa = certificado.empresa
        filial = certificado.filial

        print(f"\n[NFe Auto] Processando: {filial.nome} (CNPJ: {filial.cnpj})")

        try:
            # Verifica se certificado está vencido
            if certificado.esta_vencido:
                erro = f"Certificado vencido em {certificado.data_validade}"
                print(f"[NFe Auto] ❌ {erro}")
                config.registrar_erro(erro)
                resultados.append(f"❌ {filial.nome}: {erro}")
                continue

            # Descriptografa senha
            senha = decrypt_password(certificado.senha_encrypted)

            # Inicializa cliente SEFAZ
            client = SefazClient(
                certificado_path=certificado.arquivo_pfx.path,
                certificado_senha=senha,
                cnpj=filial.cnpj,
                uf_cod=certificado.uf_codigo
            )

            # Busca novos documentos desde último NSU
            nsu_inicial = certificado.ultimo_nsu
            print(f"[NFe Auto] Buscando desde NSU: {nsu_inicial}")

            # Faz consulta inicial
            resposta_xml = client.consultar_dfe(nsu_inicial)
            docs_temp, ult_nsu, max_nsu, mensagem = client.extrair_documentos(resposta_xml)

            # Verifica erro 656 (Consumo Indevido)
            if "Consumo Indevido" in (mensagem or ""):
                erro = "Erro 656 - Consumo Indevido da SEFAZ. Aguardando próximo ciclo."
                print(f"[NFe Auto] ⚠️ {erro}")
                # Atualiza NSU se retornado
                if ult_nsu:
                    certificado.ultimo_nsu = ult_nsu
                    certificado.save(update_fields=['ultimo_nsu'])
                config.registrar_erro(erro)
                resultados.append(f"⚠️ {filial.nome}: {erro}")
                continue

            # Busca todos os documentos
            documentos = client.buscar_todos_documentos(nsu_inicial)
            total_docs = len(documentos)
            print(f"[NFe Auto] Encontrados {total_docs} documento(s)")

            if total_docs == 0:
                print(f"[NFe Auto] ✓ Nenhum documento novo")
                # Atualiza NSU mesmo sem documentos
                if ult_nsu:
                    certificado.ultimo_nsu = ult_nsu
                    certificado.save(update_fields=['ultimo_nsu'])
                config.registrar_execucao_sucesso(0)
                resultados.append(f"✓ {filial.nome}: Nenhum documento novo")
                continue

            # Importa documentos
            importados = 0
            duplicados = 0

            with transaction.atomic():
                for xml in documentos:
                    # Verifica se é resumo e busca XML completo
                    xml_final = xml
                    if client.eh_resumo_nfe(xml):
                        chave = client.extrair_chave_resumo(xml)
                        if chave:
                            print(f"[NFe Auto] Resumo detectado, buscando completo: {chave}")
                            xml_completo = client.buscar_xml_completo(chave)
                            if xml_completo:
                                xml_final = xml_completo

                    metadados = client.extrair_metadados_nfe(xml_final)

                    # Verifica duplicata
                    if NotaFiscal.objects.filter(chave_acesso=metadados['chave_acesso']).exists():
                        duplicados += 1
                        continue

                    # Cria registro
                    nota = NotaFiscal(
                        empresa=empresa,
                        filial=filial,
                        chave_acesso=metadados['chave_acesso'],
                        numero=metadados['numero'],
                        serie=metadados['serie'],
                        data_emissao=metadados['data_emissao'],
                        emitente_cnpj=metadados['emitente_cnpj'],
                        emitente_nome=metadados['emitente_nome'],
                        valor_total=metadados['valor_total'],
                        valor_desconto=metadados['valor_desconto'],
                        valor_liquido=metadados['valor_liquido'],
                        nsu=metadados['nsu'],
                        importado_por=None  # Importação automática
                    )

                    # Salva XML
                    xml_bytes = client.xml_to_string(xml_final)
                    nota.arquivo_xml.save(
                        f"nfe_{metadados['chave_acesso']}.xml",
                        ContentFile(xml_bytes),
                        save=False
                    )

                    nota.save()
                    importados += 1

                # Atualiza último NSU
                if ult_nsu:
                    certificado.ultimo_nsu = ult_nsu
                    certificado.save(update_fields=['ultimo_nsu'])

            # Registra sucesso
            config.registrar_execucao_sucesso(importados)
            msg = f"✅ {filial.nome}: {importados} importada(s)"
            if duplicados > 0:
                msg += f" ({duplicados} duplicada(s))"
            print(f"[NFe Auto] {msg}")
            resultados.append(msg)

        except Exception as e:
            erro = f"Erro: {str(e)[:200]}"
            print(f"[NFe Auto] ❌ {filial.nome}: {erro}")
            import traceback
            traceback.print_exc()
            config.registrar_erro(erro)
            resultados.append(f"❌ {filial.nome}: {erro}")

    print(f"\n[NFe Auto] Finalizado - {timezone.now()}")
    return "\n".join(resultados)


@shared_task(name="Buscar histórico de notas fiscais")
def buscar_historico_notas():
    """
    Task que busca histórico completo de notas fiscais de forma incremental.
    Evita erro 656 fazendo pausas entre as buscas.
    """
    from financeiro.models import ConfiguracaoNFe, CertificadoDigital, NotaFiscal
    from financeiro.crypto import decrypt_password
    from financeiro.nfe.sefaz_client import SefazClient
    from django.core.files.base import ContentFile
    from django.db import transaction
    from django.utils import timezone
    import time

    print(f"[NFe Histórico] Iniciando busca histórica - {timezone.now()}")

    # Busca configurações com busca histórica ativa
    configs = ConfiguracaoNFe.objects.filter(
        busca_historica_ativa=True,
        busca_historica_status__in=['ativa', 'executando'],
        certificado__ativo=True
    ).select_related('certificado', 'certificado__filial', 'certificado__empresa')

    total_configs = configs.count()
    print(f"[NFe Histórico] Encontradas {total_configs} configuração(ões) ativa(s)")

    if total_configs == 0:
        print("[NFe Histórico] Nenhuma configuração ativa. Finalizando.")
        return "Nenhuma configuração ativa"

    resultados = []

    for config in configs:
        certificado = config.certificado
        empresa = certificado.empresa
        filial = certificado.filial

        print(f"\n[NFe Histórico] Processando: {filial.nome}")

        try:
            # Marca como executando
            config.busca_historica_status = 'executando'
            config.save()

            # Verifica se certificado está vencido
            if certificado.esta_vencido:
                erro = f"Certificado vencido em {certificado.data_validade}"
                print(f"[NFe Histórico] ❌ {erro}")
                config.busca_historica_status = 'erro'
                config.registrar_erro(erro)
                config.save()
                resultados.append(f"❌ {filial.nome}: {erro}")
                continue

            # Descriptografa senha
            senha = decrypt_password(certificado.senha_encrypted)

            # Inicializa cliente SEFAZ
            client = SefazClient(
                certificado_path=certificado.arquivo_pfx.path,
                certificado_senha=senha,
                cnpj=filial.cnpj,
                uf_cod=certificado.uf_codigo
            )

            # Busca incremental: máximo 50 documentos por execução
            # Isso evita erro 656 e permite processar gradualmente
            nsu_inicial = "000000000000000"  # Começa do início
            print(f"[NFe Histórico] Buscando desde NSU: {nsu_inicial}")

            # Consulta inicial
            resposta_xml = client.consultar_dfe(nsu_inicial)
            docs_temp, ult_nsu, max_nsu, mensagem = client.extrair_documentos(resposta_xml)

            # Verifica erro 656
            if "Consumo Indevido" in (mensagem or ""):
                erro = "Erro 656 - Aguardando próximo ciclo para continuar."
                print(f"[NFe Histórico] ⚠️ {erro}")
                config.registrar_erro(erro)
                resultados.append(f"⚠️ {filial.nome}: {erro}")
                # Mantém status como executando para tentar novamente
                continue

            # Busca documentos limitados (evita consumo excessivo)
            documentos = []
            iteracoes = 0
            max_iteracoes = 5  # Limita a 5 iterações por execução
            nsu_atual = nsu_inicial

            while iteracoes < max_iteracoes:
                iteracoes += 1
                resposta = client.consultar_dfe(nsu_atual)
                docs, novo_nsu, max_nsu_resp, msg = client.extrair_documentos(resposta)

                if not docs:
                    break

                documentos.extend(docs)
                nsu_atual = novo_nsu

                if novo_nsu == max_nsu_resp:
                    break

                # Pausa entre requisições (evita 656)
                time.sleep(2)

            total_docs = len(documentos)
            print(f"[NFe Histórico] Encontrados {total_docs} documento(s)")

            if total_docs == 0:
                print(f"[NFe Histórico] ✓ Busca histórica concluída")
                config.busca_historica_status = 'concluida'
                config.busca_historica_progresso = 100
                config.save()
                resultados.append(f"✓ {filial.nome}: Busca histórica concluída")
                continue

            # Importa documentos
            importados = 0
            duplicados = 0

            with transaction.atomic():
                for xml in documentos:
                    # Processa XML (mesmo código da busca automática)
                    xml_final = xml
                    if client.eh_resumo_nfe(xml):
                        chave = client.extrair_chave_resumo(xml)
                        if chave:
                            xml_completo = client.buscar_xml_completo(chave)
                            if xml_completo:
                                xml_final = xml_completo

                    metadados = client.extrair_metadados_nfe(xml_final)

                    if NotaFiscal.objects.filter(chave_acesso=metadados['chave_acesso']).exists():
                        duplicados += 1
                        continue

                    nota = NotaFiscal(
                        empresa=empresa,
                        filial=filial,
                        chave_acesso=metadados['chave_acesso'],
                        numero=metadados['numero'],
                        serie=metadados['serie'],
                        data_emissao=metadados['data_emissao'],
                        emitente_cnpj=metadados['emitente_cnpj'],
                        emitente_nome=metadados['emitente_nome'],
                        valor_total=metadados['valor_total'],
                        valor_desconto=metadados['valor_desconto'],
                        valor_liquido=metadados['valor_liquido'],
                        nsu=metadados['nsu'],
                        importado_por=None
                    )

                    xml_bytes = client.xml_to_string(xml_final)
                    nota.arquivo_xml.save(
                        f"nfe_{metadados['chave_acesso']}.xml",
                        ContentFile(xml_bytes),
                        save=False
                    )

                    nota.save()
                    importados += 1

            # Atualiza progresso (estimativa baseada em NSU)
            if max_nsu and ult_nsu:
                progresso = (int(ult_nsu) / int(max_nsu)) * 100
                config.busca_historica_progresso = min(int(progresso), 99)

            config.save()

            msg = f"🔄 {filial.nome}: {importados} importada(s) ({config.busca_historica_progresso}%)"
            if duplicados > 0:
                msg += f" ({duplicados} duplicada(s))"
            print(f"[NFe Histórico] {msg}")
            resultados.append(msg)

        except Exception as e:
            erro = f"Erro: {str(e)[:200]}"
            print(f"[NFe Histórico] ❌ {filial.nome}: {erro}")
            import traceback
            traceback.print_exc()
            config.busca_historica_status = 'erro'
            config.registrar_erro(erro)
            config.save()
            resultados.append(f"❌ {filial.nome}: {erro}")

    print(f"\n[NFe Histórico] Finalizado - {timezone.now()}")
    return "\n".join(resultados)
